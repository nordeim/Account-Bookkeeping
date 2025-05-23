# File: app/reporting/report_engine.py
from typing import Dict, Any, Literal, List, Optional, TYPE_CHECKING # Added TYPE_CHECKING
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle 
from reportlab.lib import colors 
from reportlab.lib.pagesizes import A4 
from reportlab.lib.units import inch, cm 
from io import BytesIO
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill 
from openpyxl.utils import get_column_letter 
# from app.core.application_core import ApplicationCore # Removed direct import
from decimal import Decimal 

if TYPE_CHECKING:
    from app.core.application_core import ApplicationCore # For type hinting

class ReportEngine:
    def __init__(self, app_core: "ApplicationCore"): # Use string literal for ApplicationCore
        self.app_core = app_core

    async def export_report(self, report_data: Dict[str, Any], format_type: Literal["pdf", "excel"]) -> bytes:
        if format_type == "pdf":
            return self._export_to_pdf_generic(report_data)
        elif format_type == "excel":
            return self._export_to_excel_generic(report_data)
        else:
            raise ValueError(f"Unsupported report format: {format_type}")

    def _format_decimal(self, value: Optional[Decimal], places: int = 2) -> str:
        if value is None: return ""
        if not isinstance(value, Decimal): 
            try:
                value = Decimal(str(value))
            except: 
                return "ERR_DEC" # Or handle as appropriate
        return f"{value:,.{places}f}"

    def _export_to_pdf_generic(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=inch, leftMargin=inch,
                                topMargin=inch, bottomMargin=inch)
        styles = getSampleStyleSheet()
        story: List[Any] = [] 

        title_text = report_data.get('title', "Financial Report")
        date_desc_text = report_data.get('report_date_description', "")
        
        story.append(Paragraph(title_text, styles['h1']))
        if date_desc_text:
            story.append(Paragraph(date_desc_text, styles['h3']))
        story.append(Spacer(1, 0.5*cm))
        
        table_data_list: List[List[Any]] = [] 

        # Trial Balance specific structure
        if 'debit_accounts' in report_data and 'credit_accounts' in report_data: 
            table_data_list = [["Account Code", "Account Name", "Debit", "Credit"]]
            for acc in report_data.get('debit_accounts', []):
                table_data_list.append([acc['code'], acc['name'], self._format_decimal(acc['balance']), ""])
            for acc in report_data.get('credit_accounts', []):
                table_data_list.append([acc['code'], acc['name'], "", self._format_decimal(acc['balance'])])
            
            totals_row = ["", Paragraph("TOTALS", styles['Heading3']), 
                         Paragraph(self._format_decimal(report_data.get('total_debits')), styles['Heading3']), 
                         Paragraph(self._format_decimal(report_data.get('total_credits')), styles['Heading3'])]
            table_data_list.append(totals_row) # type: ignore

        # Balance Sheet / P&L specific structure
        elif 'assets' in report_data or 'revenue' in report_data: 
            headers = ["Section", "Account", "Current Period"]
            if report_data.get('comparative_date') or report_data.get('comparative_start'):
                headers.append("Comparative Period")
            table_data_list = [headers]
            
            sections_map = {
                'assets': 'Assets', 'liabilities': 'Liabilities', 'equity': 'Equity',
                'revenue': 'Revenue', 'expenses': 'Expenses'
            }

            for section_key, section_title in sections_map.items():
                if section_key not in report_data: continue
                section_data = report_data.get(section_key, {})
                table_data_list.append([Paragraph(section_title, styles['h2']), ""] + (["", ""] if len(headers) == 4 else [""])) # Span section title
                
                for acc in section_data.get('accounts', []):
                    row_data = ["", f"{acc['code']} - {acc['name']}", self._format_decimal(acc['balance'])]
                    if len(headers) == 4:
                        comp_val_str = ""
                        if section_data.get('comparative_accounts'):
                            comp_acc = next((ca for ca in section_data['comparative_accounts'] if ca['id'] == acc['id']), None)
                            comp_val_str = self._format_decimal(comp_acc['balance']) if comp_acc else ""
                        row_data.append(comp_val_str)
                    table_data_list.append(row_data)
                
                total_str = self._format_decimal(section_data.get('total'))
                comp_total_str = ""
                if len(headers) == 4:
                     comp_total_str = self._format_decimal(section_data.get('comparative_total')) if section_data.get('comparative_total') is not None else ""
                
                total_row_data = ["", Paragraph(f"Total {section_title}", styles['Heading3']), total_str]
                if len(headers) == 4: total_row_data.append(comp_total_str)
                table_data_list.append(total_row_data)

            if 'net_profit' in report_data: # For P&L
                net_profit_str = self._format_decimal(report_data.get('net_profit'))
                comp_net_profit_str = ""
                if len(headers) == 4:
                    comp_net_profit_str = self._format_decimal(report_data.get('comparative_net_profit')) if report_data.get('comparative_net_profit') is not None else ""
                
                net_profit_row_data = ["", Paragraph("Net Profit/(Loss)", styles['Heading2']), net_profit_str]
                if len(headers) == 4: net_profit_row_data.append(comp_net_profit_str)
                table_data_list.append(net_profit_row_data)

        else:
            story.append(Paragraph("Report data structure not recognized for detailed PDF export.", styles['Normal']))

        if table_data_list and len(table_data_list) > 1 : # Ensure there's data beyond headers
            table = Table(table_data_list, colWidths=[doc.width/len(table_data_list[0])]*len(table_data_list[0])) 
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), 
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('ALIGN', (2,1), (-1,-1), 'RIGHT'), 
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('LEFTPADDING', (0,0), (-1,-1), 3), 
                ('RIGHTPADDING', (0,0), (-1,-1), 3),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                # Span section titles
                ('SPAN', (0,1), (1,1)), # Example for first section title if it's in row 1 (index after header)
                # This needs to be dynamic based on where section titles appear
            ]))
            story.append(table)
        
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    def _export_to_excel_generic(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active 
        
        title = report_data.get('title', "Financial Report")
        ws.title = title[:30] # type: ignore

        row_num = 1
        ws.cell(row=row_num, column=1, value=title).font = Font(bold=True, size=14) # type: ignore
        row_num += 1
        date_desc = report_data.get('report_date_description', "")
        if date_desc:
            ws.cell(row=row_num, column=1, value=date_desc).font = Font(italic=True) # type: ignore
            row_num += 1
        row_num += 1 

        if 'debit_accounts' in report_data and 'credit_accounts' in report_data: 
            headers = ["Account Code", "Account Name", "Debit", "Credit"]
            for col_num, header_text in enumerate(headers, 1): 
                ws.cell(row=row_num, column=col_num, value=header_text).font = Font(bold=True) # type: ignore
            row_num += 1

            for acc in report_data.get('debit_accounts', []):
                ws.cell(row=row_num, column=1, value=acc['code']) # type: ignore
                ws.cell(row=row_num, column=2, value=acc['name']) # type: ignore
                ws.cell(row=row_num, column=3, value=float(acc['balance'])).number_format = '#,##0.00' # type: ignore
                row_num += 1
            for acc in report_data.get('credit_accounts', []):
                ws.cell(row=row_num, column=1, value=acc['code']) # type: ignore
                ws.cell(row=row_num, column=2, value=acc['name']) # type: ignore
                ws.cell(row=row_num, column=4, value=float(acc['balance'])).number_format = '#,##0.00' # type: ignore
                row_num += 1
            
            row_num +=1 
            ws.cell(row=row_num, column=2, value="TOTALS").font = Font(bold=True) # type: ignore
            ws.cell(row=row_num, column=3, value=float(report_data.get('total_debits', Decimal(0)))).font = Font(bold=True); ws.cell(row=row_num, column=3).number_format = '#,##0.00' # type: ignore
            ws.cell(row=row_num, column=4, value=float(report_data.get('total_credits', Decimal(0)))).font = Font(bold=True); ws.cell(row=row_num, column=4).number_format = '#,##0.00' # type: ignore
        
        elif 'assets' in report_data or 'revenue' in report_data:
            headers = ["Section/Account Code", "Account Name", "Current Period"]
            has_comparative = report_data.get('comparative_date') or report_data.get('comparative_start')
            if has_comparative:
                headers.append("Comparative Period")
            
            for col_num, header_text in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=header_text).font = Font(bold=True) # type: ignore
            row_num += 1

            sections_map = {
                'assets': 'Assets', 'liabilities': 'Liabilities', 'equity': 'Equity',
                'revenue': 'Revenue', 'expenses': 'Expenses'
            }
            for section_key, section_title in sections_map.items():
                if section_key not in report_data: continue
                section_data = report_data.get(section_key, {})
                
                ws.cell(row=row_num, column=1, value=section_title).font = Font(bold=True, size=12) # type: ignore
                row_num += 1

                for acc in section_data.get('accounts', []):
                    ws.cell(row=row_num, column=1, value=acc['code']) # type: ignore
                    ws.cell(row=row_num, column=2, value=acc['name']) # type: ignore
                    ws.cell(row=row_num, column=3, value=float(acc['balance'])).number_format = '#,##0.00' # type: ignore
                    if has_comparative:
                        comp_val = ""
                        if section_data.get('comparative_accounts'):
                            comp_acc = next((ca for ca in section_data['comparative_accounts'] if ca['id'] == acc['id']), None)
                            if comp_acc and comp_acc['balance'] is not None:
                                comp_val = float(comp_acc['balance'])
                        ws.cell(row=row_num, column=4, value=comp_val).number_format = '#,##0.00' # type: ignore
                    row_num += 1
                
                ws.cell(row=row_num, column=2, value=f"Total {section_title}").font = Font(bold=True) # type: ignore
                total_val = section_data.get('total')
                ws.cell(row=row_num, column=3, value=float(total_val) if total_val is not None else "").font = Font(bold=True) # type: ignore
                ws.cell(row=row_num, column=3).number_format = '#,##0.00' # type: ignore
                if has_comparative:
                    comp_total_val = section_data.get('comparative_total')
                    ws.cell(row=row_num, column=4, value=float(comp_total_val) if comp_total_val is not None else "").font = Font(bold=True) # type: ignore
                    ws.cell(row=row_num, column=4).number_format = '#,##0.00' # type: ignore
                row_num += 1
            
            if 'net_profit' in report_data: # For P&L
                row_num += 1
                ws.cell(row=row_num, column=2, value="Net Profit/(Loss)").font = Font(bold=True, size=12) # type: ignore
                net_profit_val = report_data.get('net_profit')
                ws.cell(row=row_num, column=3, value=float(net_profit_val) if net_profit_val is not None else "").font = Font(bold=True) # type: ignore
                ws.cell(row=row_num, column=3).number_format = '#,##0.00' # type: ignore
                if has_comparative:
                    comp_net_profit_val = report_data.get('comparative_net_profit')
                    ws.cell(row=row_num, column=4, value=float(comp_net_profit_val) if comp_net_profit_val is not None else "").font = Font(bold=True) # type: ignore
                    ws.cell(row=row_num, column=4).number_format = '#,##0.00' # type: ignore

        else:
            ws.cell(row=row_num, column=1, value="Report data structure not recognized for Excel export.") # type: ignore

        for col_idx in range(1, ws.max_column + 1): # type: ignore
            column_letter = get_column_letter(col_idx)
            current_width = ws.column_dimensions[column_letter].width # type: ignore
            if current_width == 0: # openpyxl default if not set
                 ws.column_dimensions[column_letter].width = 20 # type: ignore
            else: # Auto size might make it too wide, or too narrow if default value was used.
                ws.column_dimensions[column_letter].auto_size = True # type: ignore
            
        excel_bytes_io = BytesIO()
        wb.save(excel_bytes_io)
        excel_bytes = excel_bytes_io.getvalue()
        excel_bytes_io.close()
        return excel_bytes
